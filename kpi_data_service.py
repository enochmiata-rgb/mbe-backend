from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from evidence_service import build_kpi_evidence
from market_data_service import get_brent_price


# =========================================================
# CONFIG
# =========================================================

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"

INTERNAL_KPI_SNAPSHOT_PATH = DATA_DIR / "internal_kpis.json"
INTERNAL_KPI_EXCEL_PATH = DATA_DIR / "internal_kpis.xlsx"
INTERNAL_KPI_CSV_PATH = DATA_DIR / "internal_kpis.csv"

SUPPORTED_INTERNAL_KPI_KEYS = {
    "production",
    "revenue",
    "treasury",
    "capex",
    "dividendsState",
    "headcount",
    "nationalProductionShare",
}


# =========================================================
# UTILS
# =========================================================

def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _safe_str(value: Any, default: str = "") -> str:
    if value is None:
        return default

    text = str(value).strip()
    return text if text else default


def _safe_float(value: Any, default: Optional[float] = None) -> Optional[float]:
    if value is None:
        return default

    if isinstance(value, (int, float)):
        return float(value)

    try:
        text = str(value).strip().replace(" ", "").replace(",", ".")
        return float(text)
    except Exception:
        return default


def _safe_read_json(path: Path) -> Optional[Dict[str, Any]]:
    if not path.exists():
        return None

    try:
        with open(path, "r", encoding="utf-8") as f:
            payload = json.load(f)

        if not isinstance(payload, dict):
            return None

        return payload
    except Exception as e:
        print(f"[KPI_DATA_SERVICE] JSON READ ERROR: {e}")
        return None


def _safe_write_json(path: Path, payload: Dict[str, Any]) -> bool:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"[KPI_DATA_SERVICE] JSON WRITE ERROR: {e}")
        return False


def _normalize_key(value: Any) -> str:
    raw = _safe_str(value)
    if not raw:
        return ""

    aliases = {
        "dividends_state": "dividendsState",
        "dividends-state": "dividendsState",
        "dividends state": "dividendsState",
        "dividendstate": "dividendsState",
        "dividendsstate": "dividendsState",
        "dividendesetat": "dividendsState",
        "dividendes_etat": "dividendsState",
        "dividendes état": "dividendsState",
        "national_production_share": "nationalProductionShare",
        "national-production-share": "nationalProductionShare",
        "national production share": "nationalProductionShare",
        "part production nationale": "nationalProductionShare",
        "part_production_nationale": "nationalProductionShare",
        "production": "production",
        "revenue": "revenue",
        "revenus": "revenue",
        "treasury": "treasury",
        "tresorerie": "treasury",
        "trésorerie": "treasury",
        "capex": "capex",
        "headcount": "headcount",
        "effectifs": "headcount",
    }

    compact = raw.strip()
    lowered = compact.lower()
    normalized = lowered.replace("-", "_").replace(" ", "_")

    if compact in SUPPORTED_INTERNAL_KPI_KEYS:
        return compact

    if lowered in aliases:
        return aliases[lowered]

    if normalized in aliases:
        return aliases[normalized]

    return compact


def _normalize_header(value: Any) -> str:
    text = _safe_str(value).lower()
    text = text.replace("-", "_").replace(" ", "_")
    text = text.replace("é", "e").replace("è", "e").replace("ê", "e")
    text = text.replace("à", "a").replace("ç", "c")
    return text


def _safe_int_if_possible(value: Any) -> Any:
    parsed = _safe_float(value)
    if parsed is None:
        return value

    if float(parsed).is_integer():
        return int(parsed)

    return parsed


def _find_snapshot_item(snapshot: Dict[str, Any], key: str) -> Optional[Dict[str, Any]]:
    items = snapshot.get("items", [])
    if not isinstance(items, list):
        return None

    normalized_key = _normalize_key(key)

    for item in items:
        if not isinstance(item, dict):
            continue

        if _normalize_key(item.get("key")) == normalized_key:
            return item

    return None


def _default_recommended_internal_source(key: str) -> str:
    mapping = {
        "production": (
            "Fichier de production journalière, base opérations, "
            "SCADA, ERP industriel ou consolidation terrain."
        ),
        "revenue": (
            "Grand livre, reporting finance, ERP, balance analytique "
            "ou export financier validé."
        ),
        "treasury": (
            "Position de trésorerie, cash report, ERP finance, "
            "banques consolidées ou reporting trésorerie."
        ),
        "capex": (
            "Plan CAPEX, suivi engagements, ERP projets, "
            "contrôle de gestion ou exports investissements."
        ),
        "dividendsState": (
            "Décisions CA, reporting finance, gouvernance, "
            "projection de distribution validée."
        ),
        "headcount": (
            "SIRH, paie consolidée, export RH, base effectifs "
            "ou organigramme consolidé."
        ),
        "nationalProductionShare": (
            "Production interne validée + source institutionnelle nationale "
            "pour la production pays."
        ),
    }
    return mapping.get(key, "Source métier interne à confirmer.")


def _default_unit_for_key(key: str) -> str:
    mapping = {
        "production": "bpd",
        "revenue": "xaf",
        "treasury": "xaf",
        "capex": "xaf",
        "dividendsState": "xaf",
        "headcount": "people",
        "nationalProductionShare": "percent",
    }
    return mapping.get(key, "")


def _default_provider_for_key(key: str) -> str:
    mapping = {
        "production": "Ops Control",
        "revenue": "Finance",
        "treasury": "Treasury",
        "capex": "Investments",
        "dividendsState": "Finance",
        "headcount": "HR",
        "nationalProductionShare": "Strategy",
    }
    return mapping.get(key, "Internal Source")


def _default_status_for_key(key: str) -> str:
    mapping = {
        "production": "warning",
        "revenue": "ok",
        "treasury": "warning",
        "capex": "warning",
        "dividendsState": "ok",
        "headcount": "ok",
        "nationalProductionShare": "ok",
    }
    return mapping.get(key, "ok")


def _default_title_for_key(key: str) -> str:
    mapping = {
        "production": "Production",
        "revenue": "Revenus",
        "treasury": "Trésorerie",
        "capex": "CAPEX",
        "dividendsState": "Dividendes État",
        "headcount": "Effectifs",
        "nationalProductionShare": "Part production nationale",
    }
    return mapping.get(key, key)


def _default_value_for_key(key: str) -> Any:
    mapping = {
        "production": 320000,
        "revenue": 2450000000000,
        "treasury": 98000000000,
        "capex": 420000000000,
        "dividendsState": 165000000000,
        "headcount": 2840,
        "nationalProductionShare": 31,
    }
    return mapping.get(key, 0)


def _is_valid_source_row(row: Dict[str, Any]) -> bool:
    key = _normalize_key(row.get("key"))
    if key not in SUPPORTED_INTERNAL_KPI_KEYS:
        return False

    value = _safe_float(row.get("value"))
    return value is not None


def _normalize_source_row(row: Dict[str, Any], source_path: Path) -> Optional[Dict[str, Any]]:
    key = _normalize_key(row.get("key"))

    if key not in SUPPORTED_INTERNAL_KPI_KEYS:
        return None

    value = _safe_int_if_possible(row.get("value"))
    if _safe_float(value) is None:
        return None

    title = _default_title_for_key(key)
    provider = _safe_str(row.get("provider"), _default_provider_for_key(key))
    unit = _safe_str(row.get("unit"), _default_unit_for_key(key))
    status = _safe_str(row.get("status"), _default_status_for_key(key))
    as_of = _safe_str(row.get("asOf") or row.get("as_of"), _now_iso())
    confidence = _safe_float(row.get("confidence"), 0.90) or 0.90
    evidence = _safe_str(
        row.get("evidence"),
        f"KPI {title} chargé depuis la source métier {source_path.name}.",
    )
    source_url = _safe_str(row.get("sourceUrl") or row.get("source_url"), "")

    return {
        "key": key,
        "value": value,
        "unit": unit,
        "provider": provider,
        "asOf": as_of,
        "confidence": confidence,
        "status": status,
        "evidence": evidence,
        "sourceUrl": source_url,
    }


def _read_csv_rows(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    if not path.exists():
        return rows

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as csv_file:
            reader = csv.DictReader(csv_file)
            if not reader.fieldnames:
                return rows

            for raw_row in reader:
                normalized_row: Dict[str, Any] = {}
                for raw_key, value in raw_row.items():
                    normalized_row[_normalize_header(raw_key)] = value
                rows.append(normalized_row)
    except Exception as e:
        print(f"[KPI_DATA_SERVICE] CSV READ ERROR: {e}")

    return rows


def _read_excel_rows(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    if not path.exists():
        return rows

    try:
        from openpyxl import load_workbook
    except Exception as e:
        print(f"[KPI_DATA_SERVICE] OPENPYXL UNAVAILABLE: {e}")
        return rows

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
        worksheet = workbook.active

        first_row = next(worksheet.iter_rows(min_row=1, max_row=1), None)
        if first_row is None:
            return rows

        raw_headers = [_normalize_header(cell.value) for cell in first_row]

        for excel_row in worksheet.iter_rows(min_row=2, values_only=True):
            row: Dict[str, Any] = {}
            for index, value in enumerate(excel_row):
                if index >= len(raw_headers):
                    continue

                header = raw_headers[index]
                if not header:
                    continue

                row[header] = value

            if any(value is not None and _safe_str(value) for value in row.values()):
                rows.append(row)
    except Exception as e:
        print(f"[KPI_DATA_SERVICE] EXCEL READ ERROR: {e}")

    return rows


def _source_file_to_snapshot(path: Path, source_type: str) -> Optional[Dict[str, Any]]:
    if source_type == "excel":
        rows = _read_excel_rows(path)
    elif source_type == "csv":
        rows = _read_csv_rows(path)
    else:
        rows = []

    items: List[Dict[str, Any]] = []

    for row in rows:
        if not _is_valid_source_row(row):
            continue

        normalized = _normalize_source_row(row, path)
        if normalized:
            items.append(normalized)

    if not items:
        return None

    return {
        "updatedAt": _now_iso(),
        "provider": f"internal_{source_type}",
        "items": items,
        "meta": {
            "exists": True,
            "sourceType": source_type,
            "sourcePath": str(path),
            "itemsCount": len(items),
        },
    }


def _load_internal_source_snapshot() -> Optional[Dict[str, Any]]:
    if INTERNAL_KPI_EXCEL_PATH.exists():
        snapshot = _source_file_to_snapshot(INTERNAL_KPI_EXCEL_PATH, "excel")
        if snapshot:
            _safe_write_json(INTERNAL_KPI_SNAPSHOT_PATH, snapshot)
            return snapshot

    if INTERNAL_KPI_CSV_PATH.exists():
        snapshot = _source_file_to_snapshot(INTERNAL_KPI_CSV_PATH, "csv")
        if snapshot:
            _safe_write_json(INTERNAL_KPI_SNAPSHOT_PATH, snapshot)
            return snapshot

    return None


# =========================================================
# KPI BUILDERS
# =========================================================

def _build_missing_internal_item(
    *,
    key: str,
    title: str,
    unit: str,
    default_value: Any,
    provider: str,
    status: str,
) -> Dict[str, Any]:
    recommended_source = _default_recommended_internal_source(key)
    evidence = (
        f"La source métier interne pour {title} n'est pas encore branchée. "
        f"Source recommandée : {recommended_source}"
    )

    item = build_kpi_evidence(
        key=key,
        title=title,
        value=default_value,
        unit=unit,
        provider="Internal Business Data Required",
        source_url="",
        as_of=_now_iso(),
        confidence=0.0,
        evidence=evidence,
        status=status,
        is_live=False,
        metadata={
            "fallback": True,
            "defaultProvider": provider,
            "defaultValue": default_value,
            "snapshotPath": str(INTERNAL_KPI_SNAPSHOT_PATH),
            "excelPath": str(INTERNAL_KPI_EXCEL_PATH),
            "csvPath": str(INTERNAL_KPI_CSV_PATH),
        },
    )

    item["source"] = {
        "provider": "Internal Business Data Required",
        "sourceUrl": "",
        "asOf": item["asOf"],
        "confidence": 0.0,
        "isLive": False,
        "evidence": evidence,
        "sourceMode": "internal_required",
        "sourceCategory": "internal_business_data",
    }
    item["dataCollectionStatus"] = "internal_source_required"
    item["recommendedInternalSource"] = recommended_source
    item["sourceGapEvidence"] = evidence
    item["sourceStrategy"] = {
        "mode": "internal_required",
        "path": str(INTERNAL_KPI_SNAPSHOT_PATH),
        "excelPath": str(INTERNAL_KPI_EXCEL_PATH),
        "csvPath": str(INTERNAL_KPI_CSV_PATH),
        "recommendedInternalSource": recommended_source,
    }
    return item


def _build_internal_snapshot_item(
    *,
    snapshot_item: Dict[str, Any],
    key: str,
    title: str,
    unit: str,
    provider: str,
    status: str,
) -> Dict[str, Any]:
    resolved_provider = _safe_str(snapshot_item.get("provider"), provider)
    resolved_unit = _safe_str(snapshot_item.get("unit"), unit)
    resolved_status = _safe_str(snapshot_item.get("status"), status)
    resolved_as_of = _safe_str(snapshot_item.get("asOf"), _now_iso())
    resolved_confidence = _safe_float(snapshot_item.get("confidence"), 0.85) or 0.85
    resolved_evidence = _safe_str(
        snapshot_item.get("evidence"),
        f"KPI {title} chargé depuis le snapshot métier interne.",
    )
    resolved_source_url = _safe_str(snapshot_item.get("sourceUrl"), "")
    recommended_source = _default_recommended_internal_source(key)

    item = build_kpi_evidence(
        key=key,
        title=title,
        value=snapshot_item.get("value"),
        unit=resolved_unit,
        provider=resolved_provider,
        source_url=resolved_source_url,
        as_of=resolved_as_of,
        confidence=resolved_confidence,
        evidence=resolved_evidence,
        status=resolved_status,
        is_live=False,
        metadata={
            "snapshotItem": snapshot_item,
            "snapshotPath": str(INTERNAL_KPI_SNAPSHOT_PATH),
            "excelPath": str(INTERNAL_KPI_EXCEL_PATH),
            "csvPath": str(INTERNAL_KPI_CSV_PATH),
        },
    )

    item["source"] = {
        "provider": resolved_provider,
        "sourceUrl": resolved_source_url,
        "asOf": resolved_as_of,
        "confidence": resolved_confidence,
        "isLive": False,
        "evidence": resolved_evidence,
        "sourceMode": "internal_snapshot",
        "sourceCategory": "internal_business_data",
    }
    item["dataCollectionStatus"] = "sourced"
    item["recommendedInternalSource"] = recommended_source
    item["sourceStrategy"] = {
        "mode": "internal_snapshot",
        "path": str(INTERNAL_KPI_SNAPSHOT_PATH),
        "excelPath": str(INTERNAL_KPI_EXCEL_PATH),
        "csvPath": str(INTERNAL_KPI_CSV_PATH),
        "recommendedInternalSource": recommended_source,
    }
    return item


# =========================================================
# INTERNAL SNAPSHOT
# =========================================================

def get_internal_kpi_snapshot_template() -> Dict[str, Any]:
    return {
        "updatedAt": _now_iso(),
        "provider": "internal_snapshot",
        "items": [
            {
                "key": "production",
                "value": 320000,
                "unit": "bpd",
                "provider": "Ops Control",
                "asOf": _now_iso(),
                "confidence": 0.92,
                "status": "warning",
                "evidence": "Volumes consolidés depuis la source opérations journalière.",
                "sourceUrl": "",
            },
            {
                "key": "revenue",
                "value": 2450000000000,
                "unit": "xaf",
                "provider": "Finance",
                "asOf": _now_iso(),
                "confidence": 0.90,
                "status": "ok",
                "evidence": "Revenus consolidés depuis la source finance validée.",
                "sourceUrl": "",
            },
            {
                "key": "treasury",
                "value": 98000000000,
                "unit": "xaf",
                "provider": "Treasury",
                "asOf": _now_iso(),
                "confidence": 0.89,
                "status": "warning",
                "evidence": "Position de trésorerie issue du cash report consolidé.",
                "sourceUrl": "",
            },
            {
                "key": "capex",
                "value": 420000000000,
                "unit": "xaf",
                "provider": "Investments",
                "asOf": _now_iso(),
                "confidence": 0.87,
                "status": "warning",
                "evidence": "Engagements CAPEX consolidés depuis le suivi investissements.",
                "sourceUrl": "",
            },
            {
                "key": "dividendsState",
                "value": 165000000000,
                "unit": "xaf",
                "provider": "Finance",
                "asOf": _now_iso(),
                "confidence": 0.84,
                "status": "ok",
                "evidence": "Projection de distribution validée par la finance.",
                "sourceUrl": "",
            },
            {
                "key": "headcount",
                "value": 2840,
                "unit": "people",
                "provider": "HR",
                "asOf": _now_iso(),
                "confidence": 0.96,
                "status": "ok",
                "evidence": "Effectifs consolidés depuis le SIRH.",
                "sourceUrl": "",
            },
            {
                "key": "nationalProductionShare",
                "value": 31,
                "unit": "percent",
                "provider": "Strategy",
                "asOf": _now_iso(),
                "confidence": 0.80,
                "status": "ok",
                "evidence": "Part nationale calculée depuis données internes et référence sectorielle.",
                "sourceUrl": "",
            },
        ],
    }


def get_internal_kpi_excel_template_rows() -> List[Dict[str, Any]]:
    return get_internal_kpi_snapshot_template()["items"]


def write_internal_kpi_snapshot_template() -> Dict[str, Any]:
    template = get_internal_kpi_snapshot_template()
    ok = _safe_write_json(INTERNAL_KPI_SNAPSHOT_PATH, template)

    return {
        "ok": ok,
        "path": str(INTERNAL_KPI_SNAPSHOT_PATH),
        "excelPath": str(INTERNAL_KPI_EXCEL_PATH),
        "csvPath": str(INTERNAL_KPI_CSV_PATH),
        "templateWritten": ok,
        "snapshot": template if ok else None,
    }


def load_internal_kpi_snapshot() -> Dict[str, Any]:
    source_snapshot = _load_internal_source_snapshot()
    if source_snapshot:
        items = source_snapshot.get("items", [])
        if not isinstance(items, list):
            items = []

        return {
            "updatedAt": _safe_str(source_snapshot.get("updatedAt"), ""),
            "provider": _safe_str(source_snapshot.get("provider"), "internal_source"),
            "items": [item for item in items if isinstance(item, dict)],
            "meta": {
                "exists": True,
                "path": str(INTERNAL_KPI_SNAPSHOT_PATH),
                "excelPath": str(INTERNAL_KPI_EXCEL_PATH),
                "csvPath": str(INTERNAL_KPI_CSV_PATH),
                "sourceType": source_snapshot.get("meta", {}).get("sourceType", ""),
                "sourcePath": source_snapshot.get("meta", {}).get("sourcePath", ""),
                "itemsCount": len([item for item in items if isinstance(item, dict)]),
            },
        }

    snapshot = _safe_read_json(INTERNAL_KPI_SNAPSHOT_PATH)

    if not snapshot:
        return {
            "updatedAt": "",
            "provider": "internal_snapshot_missing",
            "items": [],
            "meta": {
                "exists": False,
                "path": str(INTERNAL_KPI_SNAPSHOT_PATH),
                "excelPath": str(INTERNAL_KPI_EXCEL_PATH),
                "csvPath": str(INTERNAL_KPI_CSV_PATH),
                "itemsCount": 0,
            },
        }

    items = snapshot.get("items", [])
    if not isinstance(items, list):
        items = []

    return {
        "updatedAt": _safe_str(snapshot.get("updatedAt"), ""),
        "provider": _safe_str(snapshot.get("provider"), "internal_snapshot"),
        "items": [item for item in items if isinstance(item, dict)],
        "meta": {
            "exists": True,
            "path": str(INTERNAL_KPI_SNAPSHOT_PATH),
            "excelPath": str(INTERNAL_KPI_EXCEL_PATH),
            "csvPath": str(INTERNAL_KPI_CSV_PATH),
            "sourceType": "json",
            "sourcePath": str(INTERNAL_KPI_SNAPSHOT_PATH),
            "itemsCount": len([item for item in items if isinstance(item, dict)]),
        },
    }


# =========================================================
# KPI RESOLUTION
# =========================================================

def _resolve_brent_kpi() -> Dict[str, Any]:
    market = get_brent_price()

    price = _safe_float(market.get("price"), 85.0) or 85.0
    provider = _safe_str(market.get("provider"), "Fallback")
    source_url = _safe_str(market.get("sourceUrl"), "")
    confidence = _safe_float(market.get("confidence"), 0.55) or 0.55
    as_of = _safe_str(market.get("asOf"), _now_iso())
    evidence = _safe_str(
        market.get("evidence"),
        "Prix Brent récupéré via la source de marché configurée.",
    )
    is_live = bool(market.get("isLive", False))

    if 55 <= price <= 95:
        status = "ok"
    else:
        status = "warning"

    item = build_kpi_evidence(
        key="brent",
        title="Prix du Brent",
        value=round(price, 2),
        unit="usd",
        provider=provider,
        source_url=source_url,
        as_of=as_of,
        confidence=confidence,
        evidence=evidence,
        status=status,
        is_live=is_live,
        metadata={"marketData": market},
    )

    item["source"] = {
        "provider": provider,
        "sourceUrl": source_url,
        "asOf": as_of,
        "confidence": confidence,
        "isLive": is_live,
        "evidence": evidence,
        "sourceMode": "external_public",
        "sourceCategory": "market_data",
    }
    item["dataCollectionStatus"] = "sourced"
    item["recommendedInternalSource"] = ""
    item["sourceStrategy"] = {
        "mode": "external_public",
        "path": "",
        "recommendedInternalSource": "",
    }
    return item


def _resolve_internal_kpi(
    *,
    key: str,
    title: str,
    unit: str,
    default_value: Any,
    provider: str,
    status: str,
    snapshot: Dict[str, Any],
) -> Dict[str, Any]:
    snapshot_item = _find_snapshot_item(snapshot, key)

    if snapshot_item:
        return _build_internal_snapshot_item(
            snapshot_item=snapshot_item,
            key=key,
            title=title,
            unit=unit,
            provider=provider,
            status=status,
        )

    return _build_missing_internal_item(
        key=key,
        title=title,
        unit=unit,
        default_value=default_value,
        provider=provider,
        status=status,
    )


def resolve_core_kpis() -> List[Dict[str, Any]]:
    snapshot = load_internal_kpi_snapshot()

    items: List[Dict[str, Any]] = [
        _resolve_brent_kpi(),
        _resolve_internal_kpi(
            key="production",
            title="Production",
            unit="bpd",
            default_value=320000,
            provider="Ops Control",
            status="warning",
            snapshot=snapshot,
        ),
        _resolve_internal_kpi(
            key="revenue",
            title="Revenus",
            unit="xaf",
            default_value=2450000000000,
            provider="Finance",
            status="ok",
            snapshot=snapshot,
        ),
        _resolve_internal_kpi(
            key="treasury",
            title="Trésorerie",
            unit="xaf",
            default_value=98000000000,
            provider="Treasury",
            status="warning",
            snapshot=snapshot,
        ),
        _resolve_internal_kpi(
            key="capex",
            title="CAPEX",
            unit="xaf",
            default_value=420000000000,
            provider="Investments",
            status="warning",
            snapshot=snapshot,
        ),
        _resolve_internal_kpi(
            key="dividendsState",
            title="Dividendes État",
            unit="xaf",
            default_value=165000000000,
            provider="Finance",
            status="ok",
            snapshot=snapshot,
        ),
        _resolve_internal_kpi(
            key="headcount",
            title="Effectifs",
            unit="people",
            default_value=2840,
            provider="HR",
            status="ok",
            snapshot=snapshot,
        ),
        _resolve_internal_kpi(
            key="nationalProductionShare",
            title="Part production nationale",
            unit="percent",
            default_value=31,
            provider="Strategy",
            status="ok",
            snapshot=snapshot,
        ),
    ]

    return items
